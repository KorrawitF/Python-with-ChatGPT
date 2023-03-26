import openai
import openpyxl
import time

openai.api_key = "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" #initialize api key use your api from 
file = 'Copy of Chat Gpt test.xlsx' #choose xlsx file to use
theFile = openpyxl.load_workbook(file) 
currentSheet = theFile['Sheet1'] #choose sheet to query
len = len(currentSheet['A'])-1 #Have to minus 1 to avoid first row which is header.
#Prepare chatgpt to get the answer
messages = [
    {"role": "system", 
     "content":  "I will give you a company name and a description. I want you to guess if it is a gym academy. Answer it with the following options only: High likely, Medium likely, Low likely, no, if the description of any of them is None just reply as 'Need more info'. No need any comment."
     } 
]
  
# Get information from xlsx file and give chatgpt.
for col in range(len):
    try:
        if currentSheet.cell(row=(col+2), column=4).value == None:
            message = "Name: " + currentSheet.cell(row=(col+2), column=1).value + ", Description: " + str(currentSheet.cell(row=(col+2), column=2).value) #Declare chatgpt information Name: name, Description: First desc.
            if messages:
                messages.append(
                            {"role": "user", "content": message},
                    )
            chat_completion = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=messages
            )
            answer = chat_completion.choices[0].message.content
            currentSheet.cell(row=(col+2), column=4).value = f"{answer}"
            messages.append({"role": "assistant", "content": answer})
            theFile.save(file)
            print(currentSheet.cell(row=(col+2), column=1).value + ', Done!')
            time.sleep(1.5) #Delay to avoid request limit
        else:
            continue
    except: #catch the limit error and wait for a while then continue.
        time.sleep(5)
        continue